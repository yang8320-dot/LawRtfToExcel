import os
import re
import ctypes
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinterdnd2 import TkinterDnD, DND_FILES  # 匯入拖曳套件
from datetime import datetime
from striprtf.striprtf import rtf_to_text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================
# 解決 Windows 125% / 150% 縮放導致文字模糊
# ==========================================
if os.name == 'nt':
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

# ==========================================
# 核心邏輯區塊 (維持不變)
# ==========================================
def chinese_to_arabic(chn_str: str) -> int:
    s = chn_str.replace("、", "").replace("(", "").replace(")", "")
    s = s.replace("（", "").replace("）", "").strip()
    if s.isdigit(): return int(s)
    num_dict = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, 
                "六": 6, "七": 7, "八": 8, "九": 9, "十": 10,
                "百": 100, "千": 1000, "零": 0}
    total, current = 0, 0
    for char in s:
        if char in num_dict:
            v = num_dict[char]
            if v in [10, 100, 1000]:
                if current == 0: current = 1
                total += current * v
                current = 0
            else:
                current = v
    total += current
    return total

def format_tiao(article_str: str) -> str:
    pure = article_str.replace("第", "").replace("條", "").replace("之", "-").strip()
    parts = pure.split("-")
    if len(parts) > 0:
        main_num = chinese_to_arabic(parts[0])
        if len(parts) > 1:
            sub_num = chinese_to_arabic(parts[1])
            return f"{main_num:03d}-{sub_num}"
        else:
            return f"{main_num:03d}"
    return article_str

def get_rtf_text(file_path: str) -> str:
    try:
        try:
            with open(file_path, 'r', encoding='cp950') as file: rtf_content = file.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file: rtf_content = file.read()
        return rtf_to_text(rtf_content)
    except Exception as e:
        raise Exception(f"讀取 RTF 失敗：{str(e)}")

def write_to_excel(records, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "法規轉換資料"
    
    headers = ["日期", "法令名稱", "條", "項", "款", "目", "內容", "重點摘要", 
               "適用性", "符合度", "有提升績效機會", "有潛在不符合風險", "鑑別日期", "備註"]
    header_fill = PatternFill(start_color="FFEE99", end_color="FFEE99", fill_type="solid")
    header_font = Font(name="微軟正黑體", bold=True)
    
    ws.append(headers)
    for col_idx in range(1, 15):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_data in records: ws.append(row_data)
        
    max_row = len(records) + 1
    font_default = Font(name="微軟正黑體")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    col_widths = {'A': 12, 'B': 20, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 55, 
                  'H': 40, 'I': 10, 'J': 10, 'K': 15, 'L': 15, 'M': 12, 'N': 40}
    for col, width in col_widths.items(): ws.column_dimensions[col].width = width

    for r in range(2, max_row + 1):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c); cell.font = font_default
            if c in [3, 4, 5, 6]: cell.number_format = '@'; cell.alignment = align_center
            elif c in [2, 7, 8, 14]: cell.alignment = align_left_wrap
            elif c in [1, 9, 10, 11, 12, 13]: cell.alignment = align_center
                
    ws.freeze_panes = "A2"
    if max_row >= 2:
        dv_i = DataValidation(type="list", formula1='"適用,不適用,參考,確認中"', showDropDown=True)
        dv_j = DataValidation(type="list", formula1='"合法,不合法"', showDropDown=True)
        dv_kl = DataValidation(type="list", formula1='" ,v"', showDropDown=True)
        ws.add_data_validation(dv_i); ws.add_data_validation(dv_j); ws.add_data_validation(dv_kl)
        dv_i.add(f"I2:I{max_row}"); dv_j.add(f"J2:J{max_row}"); dv_kl.add(f"K2:K{max_row}"); dv_kl.add(f"L2:L{max_row}")

    wb.save(save_path)

def process_law_data(rtf_path: str, save_path: str, stop_event: threading.Event) -> bool:
    plain_text = get_rtf_text(rtf_path)
    if not plain_text: return False

    lines = plain_text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    law_name, law_date = "", ""
    current_article, current_xiang_num = "", 0
    current_item_lv3, current_item_lv4, current_content = "", "", ""
    
    reg_date = re.compile(r"(修正|發布)日期：?\s*民國\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日")
    reg_article = re.compile(r"^第\s*([一二三四五六七八九十百千\d]+)\s*(?:[-之]\s*([一二三四五六七八九十百千\d]+)\s*)?條(?:[-之]\s*([一二三四五六七八九十百千\d]+))?")
    reg_kuan = re.compile(r"^[一二三四五六七八九十百千]+、")
    reg_mu = re.compile(r"^[（\(][一二三四五六極七八九十百千]+[）\)]")
    reg_trim_digits = re.compile(r"^\d+\s+")
    
    records = []
    def add_record():
        if current_article and current_content.strip():
            records.append([law_date, law_name, current_article, f"{current_xiang_num:02d}", 
                            current_item_lv3, current_item_lv4, current_content.strip(), 
                            "", "", "", "", "", datetime.now().strftime("%Y-%m-%d"), ""])

    for i, line in enumerate(lines):
        if i % 50 == 0 and stop_event.is_set(): return False
        line = line.strip()
        line = reg_trim_digits.sub("", line)
        if not line: continue
            
        if not law_name:
            if line.startswith("法規名稱："): law_name = line[5:].strip(); continue
            elif not line.startswith("第") and "日期" not in line and len(line) < 50: law_name = line; continue
                
        if not law_date:
            match = reg_date.search(line)
            if match:
                law_date = f"{int(match.group(2)) + 1911}-{int(match.group(3)):02d}-{int(match.group(4)):02d}"
                continue
                
        match_article = reg_article.search(line)
        if match_article:
            add_record()
            match_val = match_article.group(0)
            current_article = format_tiao(match_val)
            current_xiang_num, current_item_lv3, current_item_lv4 = 1, "", ""
            current_content = line[len(match_val):].strip()
            continue
            
        if current_article:
            match_kuan = reg_kuan.search(line)
            match_mu = reg_mu.search(line)
            if match_kuan:
                add_record()
                current_item_lv3 = f"{chinese_to_arabic(match_kuan.group(0)):02d}"
                current_item_lv4, current_content = "", line[len(match_kuan.group(0)):].strip()
            elif match_mu:
                add_record()
                current_item_lv4 = f"{chinese_to_arabic(match_mu.group(0)):02d}"
                current_content = line[len(match_mu.group(0)):].strip()
            else:
                if not current_item_lv3 and not current_item_lv4:
                    if current_content.strip() and current_content.strip()[-1] in ["。", "：", ":", "."]:
                        add_record()
                        current_content, current_xiang_num = "", current_xiang_num + 1
                current_content = f"{current_content}\n{line}" if current_content else line

    add_record()
    if stop_event.is_set(): return False
    write_to_excel(records, save_path)
    return True

# ==========================================
# GUI 介面區塊 (包含拖曳與刪除功能)
# ==========================================
class LawConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("法規 RTF 轉 Excel 批次工具")
        self.root.geometry("750x500")
        self.root.option_add("*Font", ("微軟正黑體", 11))
        
        self.file_paths = []
        self.stop_event = threading.Event()
        self.setup_ui()

    def setup_ui(self):
        style = ttk.Style()
        style.configure("TButton", font=("微軟正黑體", 11))
        
        # 頂部按鈕區
        frame_top = tk.Frame(self.root, pady=12, padx=15)
        frame_top.pack(fill=tk.X)
        ttk.Button(frame_top, text="加入檔案 (可多選)", command=self.add_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_top, text="清空清單", command=self.clear_files).pack(side=tk.LEFT, padx=5)
        tk.Label(frame_top, text="💡 提示：支援拖曳檔案、Delete鍵/右鍵可刪除", fg="gray").pack(side=tk.RIGHT, padx=5)

        # 中間清單區 (加入拖曳設定)
        frame_mid = tk.Frame(self.root, padx=20)
        frame_mid.pack(fill=tk.BOTH, expand=True)
        
        self.listbox = tk.Listbox(frame_mid, selectmode=tk.EXTENDED, font=("微軟正黑體", 11))
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(frame_mid, orient=tk.VERTICAL, command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scrollbar.set)

        # 🎯 註冊拖曳事件
        self.listbox.drop_target_register(DND_FILES)
        self.listbox.dnd_bind('<<Drop>>', self.on_file_drop)

        # 🎯 註冊刪除按鍵 (Delete)
        self.listbox.bind('<Delete>', self.delete_selected)
        
        # 🎯 註冊右鍵選單
        self.context_menu = tk.Menu(self.root, tearoff=0, font=("微軟正黑體", 10))
        self.context_menu.add_command(label="🗑️ 刪除選取檔案", command=self.delete_selected)
        self.listbox.bind('<Button-3>', self.show_context_menu)

        # 底部控制區
        frame_bottom = tk.Frame(self.root, pady=15, padx=15)
        frame_bottom.pack(fill=tk.X)

        self.btn_run = ttk.Button(frame_bottom, text="開始執行轉換", command=self.start_processing)
        self.btn_run.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        self.btn_stop = ttk.Button(frame_bottom, text="停止 (中止)", command=self.stop_processing, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT, padx=10)

        # 狀態與進度條
        self.lbl_status = tk.Label(self.root, text="準備就緒", fg="blue", font=("微軟正黑體", 10, "bold"))
        self.lbl_status.pack(pady=5)
        self.progress = ttk.Progressbar(self.root, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.pack(fill=tk.X, padx=20, pady=10)

    # 拖曳檔案處理邏輯
    def on_file_drop(self, event):
        # 將 Tkinter 的拖曳字串轉換為 Python 的清單格式
        files = self.root.tk.splitlist(event.data)
        added_count = 0
        for f in files:
            if f.lower().endswith(".rtf") and f not in self.file_paths:
                self.file_paths.append(f)
                self.listbox.insert(tk.END, f)
                added_count += 1
                
        if added_count > 0:
            self.lbl_status.config(text=f"已透過拖曳加入 {added_count} 個檔案，共 {len(self.file_paths)} 個")
        else:
            self.lbl_status.config(text="未加入任何新檔案 (僅支援 .rtf 格式或檔案已存在)", fg="red")

    # 刪除選取檔案邏輯
    def delete_selected(self, event=None):
        selected_indices = self.listbox.curselection()
        if not selected_indices:
            return
            
        # 必須從後面開始刪除，才不會影響前面的 index
        for i in reversed(selected_indices):
            self.listbox.delete(i)
            del self.file_paths[i]
            
        self.lbl_status.config(text=f"已刪除選取檔案，清單剩餘 {len(self.file_paths)} 個檔案", fg="blue")

    # 顯示右鍵選單
    def show_context_menu(self, event):
        # 判斷右鍵點擊的位置是否有項目
        clicked_index = self.listbox.nearest(event.y)
        
        # 若點擊的項目不在目前的選取範圍內，就單獨選取該項目
        if clicked_index not in self.listbox.curselection():
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(clicked_index)
            self.listbox.activate(clicked_index)
            
        # 只要有選取的項目就顯示右鍵選單
        if self.listbox.curselection():
            self.context_menu.post(event.x_root, event.y_root)

    def add_files(self):
        files = filedialog.askopenfilenames(title="選擇法規 RTF 檔案", filetypes=[("RTF Files", "*.rtf")])
        for f in files:
            if f not in self.file_paths:
                self.file_paths.append(f)
                self.listbox.insert(tk.END, f)
        self.lbl_status.config(text=f"目前清單共 {len(self.file_paths)} 個檔案", fg="blue")

    def clear_files(self):
        self.file_paths.clear()
        self.listbox.delete(0, tk.END)
        self.lbl_status.config(text="清單已清空")

    def stop_processing(self):
        self.stop_event.set()
        self.lbl_status.config(text="正在停止中...請稍候", fg="red")
        self.btn_stop.config(state=tk.DISABLED)

    def set_ui_state(self, is_running):
        state = tk.DISABLED if is_running else tk.NORMAL
        self.btn_run.config(state=state)
        self.btn_stop.config(state=tk.NORMAL if is_running else tk.DISABLED)

    def start_processing(self):
        if not self.file_paths:
            messagebox.showwarning("警告", "請先加入至少一個 RTF 檔案！")
            return
            
        self.set_ui_state(is_running=True)
        self.stop_event.clear()
        self.progress['maximum'] = len(self.file_paths)
        self.progress['value'] = 0
        self.lbl_status.config(text="開始處理...", fg="blue")
        
        thread = threading.Thread(target=self.process_files_thread)
        thread.daemon = True
        thread.start()

    def process_files_thread(self):
        success_count = 0
        error_count = 0
        
        for i, file_path in enumerate(self.file_paths):
            if self.stop_event.is_set():
                self.update_status(f"處理已手動中斷！(完成 {success_count} 筆)", "red")
                break
                
            self.update_status(f"正在處理 ({i+1}/{len(self.file_paths)}): {os.path.basename(file_path)}", "blue")
            save_path = file_path.replace(".rtf", "_轉換結果.xlsx")
            
            try:
                is_completed = process_law_data(file_path, save_path, self.stop_event)
                if is_completed:
                    success_count += 1
            except Exception as e:
                print(e)
                error_count += 1
                
            self.root.after(0, self.progress.step, 1)

        if not self.stop_event.is_set():
            self.update_status(f"全部完成！成功: {success_count}，失敗: {error_count}", "green")
            self.root.after(0, lambda: messagebox.showinfo("完成", f"處理完畢！\n成功: {success_count}\n失敗: {error_count}\n(產生的 Excel 檔案位於原 RTF 檔案同目錄下)"))
            
        self.root.after(0, lambda: self.set_ui_state(is_running=False))

    def update_status(self, text, color):
        self.root.after(0, lambda: self.lbl_status.config(text=text, fg=color))

if __name__ == "__main__":
    # 使用支援拖曳的 TkinterDnD 根視窗
    root = TkinterDnD.Tk()
    app = LawConverterApp(root)
    root.mainloop()
